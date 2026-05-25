import os
import subprocess
import webbrowser
import shutil
from turtle import title

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLineEdit, QLabel, QPushButton,
    QComboBox, QTextEdit, QFileDialog, QMessageBox, QGroupBox, QScrollArea, QHBoxLayout
)
from PyQt5.QtGui import QFont, QIcon
import sqlite3
from openpyxl import Workbook, load_workbook
import os

from whatsapp.whatsapp_menu import WhatsAppMenu

class OrdersForm(QWidget):
    def __init__(self, parent=None):   # parent = Dashboard
        super().__init__(parent)
        self.setWindowTitle("New Order")
        self.resize(700, 800)
        self.setFont(QFont("Roboto", 11))

        main_layout = QVBoxLayout()

        # --- Dashboard Menu Link ---
        top_bar = QHBoxLayout()
        back_btn = QPushButton("← Back to Dashboard")
        back_btn.setStyleSheet("QPushButton { background: #444; color: white; padding: 4px 8px; border-radius: 4px; }")
        back_btn.clicked.connect(self.close)
        top_bar.addWidget(back_btn)
        main_layout.addLayout(top_bar)

        # Scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        layout = QVBoxLayout(content)

        # --- A. Basic Details ---
        group_basic = QGroupBox("Basic Details")
        basic_layout = QVBoxLayout()
        self.date = QLineEdit(); 
        self.date.setPlaceholderText("Date (DD-MM-YYYY)")
        self.broker_name = QLineEdit(); 
        self.broker_name.setPlaceholderText("Broker Name (optional)")
        group_basic = self.make_card("Basic Details", [self.date, self.broker_name])
        layout.addWidget(group_basic)

        # --- B. Vehicle Details ---
        group_vehicle = QGroupBox("Vehicle Details")
        vehicle_layout = QVBoxLayout()
        self.vehicle_no = QLineEdit()
        self.vehicle_no.setPlaceholderText("Vehicle Number (Lorry No)")
        self.vehicle_type = QLineEdit()
        self.vehicle_type.setPlaceholderText("Vehicle Type")
        self.owner_name = QLineEdit()
        self.owner_name.setPlaceholderText("Owner Name")
        self.driver_name = QLineEdit()
        self.driver_name.setPlaceholderText("Driver Name")
        self.driver_mobile = QLineEdit()
        self.driver_mobile.setPlaceholderText("Driver Mobile Number")
        for field in [self.vehicle_no, self.vehicle_type, self.owner_name, self.driver_name, self.driver_mobile]:
            vehicle_layout.addWidget(field)
        group_vehicle.setLayout(vehicle_layout)
        layout.addWidget(group_vehicle)

        # --- C. Trip Details ---
        group_trip = QGroupBox("Trip Details")
        trip_layout = QVBoxLayout()
        self.from_location = QLineEdit()
        self.from_location.setPlaceholderText("From Location")
        self.to_location = QLineEdit()
        self.to_location.setPlaceholderText("To Location")
        self.material = QLineEdit()
        self.material.setPlaceholderText("Material")
        self.load_weight = QLineEdit()
        self.load_weight.setPlaceholderText("Load Weight (tons)")
        for field in [self.from_location, self.to_location, self.material, self.load_weight]:
            trip_layout.addWidget(field)
        group_trip.setLayout(trip_layout)
        layout.addWidget(group_trip)

        # --- D. Payment Details ---
        group_payment = QGroupBox("Payment Details")
        payment_layout = QVBoxLayout()
        self.freight_amount = QLineEdit(); self.freight_amount.setPlaceholderText("Freight Amount")
        self.advance_paid = QLineEdit(); self.advance_paid.setPlaceholderText("Advance Paid")
        self.balance_amount = QLineEdit(); self.balance_amount.setPlaceholderText("Balance (auto)")
        self.balance_amount.setReadOnly(True)
        group_payment = self.make_card("Payment Details", [self.freight_amount, self.advance_paid, self.balance_amount])
        layout.addWidget(group_payment)

        # --- E. Broker Commission ---
        group_commission = QGroupBox("Broker Commission")
        commission_layout = QVBoxLayout()
        self.commission_type = QLineEdit()
        self.commission_type.setPlaceholderText("Commission Type")
        self.commission_amount = QLineEdit()
        self.commission_amount.setPlaceholderText("Commission Amount")
        commission_layout.addWidget(self.commission_type)
        commission_layout.addWidget(self.commission_amount)
        group_commission.setLayout(commission_layout)
        layout.addWidget(group_commission)

        # --- F. Status Tracking ---
        group_status = QGroupBox("Status Tracking")
        status_layout = QVBoxLayout()
        self.order_status = QComboBox()
        self.order_status.addItems(["Pending", "In Transit", "Completed"])
        self.payment_status = QComboBox()
        self.payment_status.addItems(["Unpaid", "Partial", "Paid"])
        status_layout.addWidget(self.order_status)
        status_layout.addWidget(self.payment_status)
        group_status.setLayout(status_layout)
        layout.addWidget(group_status)

        # --- G. Extras ---
        group_extras = QGroupBox("Extras")
        extras_layout = QVBoxLayout()
        self.remark = QTextEdit()
        self.remark.setPlaceholderText("Remarks / Notes")
        self.bill_btn = QPushButton("📎 Upload Bill/LR")
        self.bill_btn.clicked.connect(self.upload_bill)
        extras_layout.addWidget(self.remark)
        extras_layout.addWidget(self.bill_btn)
        group_extras.setLayout(extras_layout)
        layout.addWidget(group_extras)


        # Save button
        save_btn = QPushButton("💾 Save Order")
        save_btn.clicked.connect(self.save_order)         # ✅ will work once save_order is defined
        layout.addWidget(save_btn)

        # Attach form to scroll
        scroll.setWidget(content)
        main_layout.addWidget(scroll)
        self.setLayout(main_layout)

        self.bill_path = None

        #Whatsapp button in top bar
        self.whatsapp_btn = QPushButton(" WhatsApp")
        self.whatsapp_btn.setIcon(QIcon("whatsapp/icons/whatsapp.png"))
        self.whatsapp_btn.setStyleSheet("""
            QPushButton {
                background-color: #25D366;
                color: white;
                font-weight: bold;
                padding: 8px 12px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #128C7E;
            }
        """)
        self.whatsapp_btn.clicked.connect(self.open_whatsapp_menu)
        top_bar.addWidget(self.whatsapp_btn)
        # --- Enter key navigation ---
        self.date.returnPressed.connect(self.handle_date_enter)
        self.broker_name.returnPressed.connect(lambda: self.vehicle_no.setFocus())
        self.vehicle_no.returnPressed.connect(lambda: self.vehicle_type.setFocus())
        self.vehicle_type.returnPressed.connect(lambda: self.owner_name.setFocus())
        self.owner_name.returnPressed.connect(lambda: self.driver_name.setFocus())
        self.driver_name.returnPressed.connect(lambda: self.driver_mobile.setFocus())
        self.driver_mobile.returnPressed.connect(lambda: self.from_location.setFocus())
        self.from_location.returnPressed.connect(lambda: self.to_location.setFocus())
        self.to_location.returnPressed.connect(lambda: self.material.setFocus())
        self.material.returnPressed.connect(lambda: self.load_weight.setFocus())
        self.load_weight.returnPressed.connect(lambda: self.freight_amount.setFocus())
        self.freight_amount.returnPressed.connect(lambda: self.advance_paid.setFocus())
        self.advance_paid.returnPressed.connect(lambda: self.balance_amount.setFocus())
        self.balance_amount.returnPressed.connect(lambda: self.commission_type.setFocus())
        self.commission_type.returnPressed.connect(lambda: self.commission_amount.setFocus())
        self.commission_amount.returnPressed.connect(lambda: self.order_status.setFocus())
        self.order_status.activated.connect(lambda: self.payment_status.setFocus())
        self.payment_status.activated.connect(lambda: self.remark.setFocus())

        # Auto-update balance
        self.freight_amount.textChanged.connect(self.update_balance)   # ✅ will work once update_balance is defined
        self.advance_paid.textChanged.connect(self.update_balance)

        # --- Style ---
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 1px solid #ccc;
                border-radius: 6px;
                margin-top: 10px;
                padding: 6px;
            }
            QLineEdit, QTextEdit, QComboBox {
                padding: 6px;
                border: 1px solid #ccc;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton {
                background-color: #1976D2;
                color: white;
                padding: 8px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1565C0;
            }
        """)

    def make_card(self, title, widgets):
        card = QGroupBox(title)
        layout = QVBoxLayout()
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(8)
        for w in widgets:
            layout.addWidget(w)
        card.setLayout(layout)

        card.setStyleSheet("""
            QGroupBox {
                background: #fff;
                border: 1px solid #d0d7de;
                border-radius: 8px;
                margin-top: 10px;
                font-weight: bold;
                padding-top: 8px;
            }
        """)
        return card

    # --- Helpers ---
    def format_date(self):
        text = self.date.text().strip()
        # If user types 8 digits like 29122026
        if len(text) == 8 and text.isdigit():
            day = text[0:2]
            month = text[2:4]
            year = text[4:]
            formatted = f"{day}/{month}/{year}"
            self.date.setText(formatted)

    def handle_date_enter(self):
        self.format_date()
        self.broker_name.setFocus()
    import subprocess, webbrowser

    def open_whatsapp_menu(self):
        try:
            # 1. Check if standalone WhatsApp.exe exists
            exe_path = r"C:\Users\{}\AppData\Local\WhatsApp\WhatsApp.exe".format(os.getlogin())
            if os.path.exists(exe_path):
                subprocess.Popen([exe_path])
                return

            # 2. Try Microsoft Store version (URI scheme)
            result = os.system("start whatsapp://")
            if result == 0:
                return

            # 3. Fallback → WhatsApp Web
            webbrowser.open("https://web.whatsapp.com/")
        except Exception as e:
            print("Error opening WhatsApp:", e)
            # Always fallback to Web if anything fails
            webbrowser.open("https://web.whatsapp.com/")


    def update_balance(self):
        try:
            freight = float(self.freight_amount.text() or 0)
            advance = float(self.advance_paid.text() or 0)
            balance = freight - advance
            self.balance_amount.setText(str(balance))
        except ValueError:
            self.balance_amount.setText("")

    def upload_bill(self):
        path, _ = QFileDialog.getOpenFileName(self, "Upload Bill/LR")
        if path:
            self.bill_path = path

    def save_order(self):
        # --- Validation ---
        if not self.date.text().strip():
            QMessageBox.warning(self, "Missing Data", "Date is required.")
            return
        if not self.vehicle_no.text().strip():
            QMessageBox.warning(self, "Missing Data", "Vehicle Number is required.")
            return
        if not self.from_location.text().strip() or not self.to_location.text().strip():
            QMessageBox.warning(self, "Missing Data", "Trip locations are required.")
            return
        if not self.freight_amount.text().strip():
            QMessageBox.warning(self, "Missing Data", "Freight Amount is required.")
            return

        # Balance calculation
        freight = float(self.freight_amount.text() or 0)
        advance = float(self.advance_paid.text() or 0)
        balance = freight - advance
        self.balance_amount.setText(str(balance))

        # --- Save into golden_transport.db ---
        conn = sqlite3.connect("golden_transport.db")
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO orders (date, lorry_no, add_name, driver_name,
                                from_place, to_place, freight, ton_age,
                                advance, balance, load, commission, lorry_name)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            self.date.text(),
            self.vehicle_no.text(),
            self.broker_name.text(),
            self.driver_name.text(),
            self.from_location.text(),
            self.to_location.text(),
            freight,
            self.load_weight.text(),
            advance,
            balance,
            self.material.text(),
            self.commission_amount.text(),
            self.vehicle_type.text()
        ))
        conn.commit()
        conn.close()

        # --- Save into Excel ---
        excel_file = "orders.xlsx"

        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Header row
            ws.append(["Date","Lorry No","Broker Name","Driver Name","From","To",
                       "Freight","Ton_Age","Advance","Balance","Load","Commission","Lorry Name"])

        # Append new order row
        ws.append([
            self.date.text(),
            self.vehicle_no.text(),
            self.broker_name.text(),
            self.driver_name.text(),
            self.from_location.text(),
            self.to_location.text(),
            freight,
            self.load_weight.text(),
            advance,
            balance,
            self.material.text(),
            self.commission_amount.text(),
            self.vehicle_type.text()
        ])

        wb.save(excel_file)

        QMessageBox.information(self, "Success", "Order saved to DB and Excel.")
        self.close()
